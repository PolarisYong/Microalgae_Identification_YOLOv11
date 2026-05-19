import io
import sys

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
from matplotlib import rcParams
from matplotlib.ticker import MultipleLocator
from openpyxl.drawing.image import Image
from openpyxl.styles import numbers
from scipy.optimize import curve_fit

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")

# ---------------------- 全局字体配置 ----------------------
rcParams["font.family"] = ["Times New Roman", "serif"]
rcParams["axes.unicode_minus"] = False
rcParams["font.size"] = 10
rcParams["axes.labelsize"] = 20
rcParams["xtick.labelsize"] = 20
rcParams["ytick.labelsize"] = 20
rcParams["axes.titlesize"] = 24
rcParams["legend.fontsize"] = 18
rcParams["axes.titley"] = 1.01


# ---------------------- 1. 定义修正Logistic生长模型 ----------------------
def modified_logistic_model(t, mu_max, K, N0):
    t = np.asarray(t, dtype=float)
    if K == 0 or N0 == 0:
        return np.zeros_like(t, dtype=float)
    term = (K / N0 - 1) * np.exp(-mu_max * t)
    N_t = K / (1 + term)
    return N_t


def calculate_growth_phases(t, mu_max, K, N0):
    t = np.asarray(t, dtype=float)
    if t.size == 0:
        return {
            "滞后期时长(h)": 0,
            "对数期时长(h)": 0,
            "稳定期时长(h)": 0,
        }, []

    term = (K / N0 - 1) * np.exp(-mu_max * t)
    N_t = K / (1 + term)

    dt = np.diff(t).mean() if len(t) > 1 else 1
    if dt == 0:
        dt = 1
    slope = np.diff(N_t) / dt
    slope = np.insert(slope, 0, 0)

    max_slope = np.max(slope) if np.max(slope) > 0 else 1e-6
    threshold = 0.2 * max_slope

    phase_flags = []
    for i in range(len(t)):
        current_slope = slope[i]
        current_N = N_t[i]
        if K != 0 and current_slope < threshold and abs(current_N - K) / K < 0.1:
            phase_flags.append("稳定期")
        elif current_slope >= threshold:
            phase_flags.append("对数期")
        else:
            phase_flags.append("滞后期")

    phases = []
    current_phase = phase_flags[0]
    start_idx = 0
    for i in range(1, len(phase_flags)):
        if phase_flags[i] != current_phase:
            start_time = t[start_idx]
            # 边界归后段：前一阶段结束在下一阶段起点，保证总时长不丢失边界小时
            end_time = t[i]
            duration = end_time - start_time
            phases.append({
                "阶段": current_phase,
                "开始时间(h)": start_time,
                "结束时间(h)": end_time,
                "时长(h)": duration,
            })
            current_phase = phase_flags[i]
            start_idx = i

    start_time = t[start_idx]
    end_time = t[-1]
    duration = end_time - start_time
    phases.append({
        "阶段": current_phase,
        "开始时间(h)": start_time,
        "结束时间(h)": end_time,
        "时长(h)": duration,
    })

    phase_duration = {
        "滞后期时长(h)": 0,
        "对数期时长(h)": 0,
        "稳定期时长(h)": 0,
    }
    for phase in phases:
        key = f"{phase['阶段']}时长(h)"
        if key in phase_duration:
            phase_duration[key] = round(max(phase["时长(h)"], 0), 2)
    return phase_duration, phases


# ---------------------- 2. 生成趋势线函数 ----------------------
def generate_trendline(x, y, degree=2):
    z = np.polyfit(x, y, degree)
    p = np.poly1d(z)
    return p(x)


def _empty_summary_row(sheet_name):
    return {
        "腔室名称": sheet_name,
        "最大比生长速率μmax (h^-1)": np.nan,
        "环境容纳量K (个/腔室)": np.nan,
        "初始细胞数量N0 (个/腔室)": np.nan,
        "拟合优度R²": np.nan,
        "平均细胞周期T_d(h)": np.nan,
        "增殖倍数 F": np.nan,
        "生长效率 η (1/增殖倍数)": np.nan,
        "滞后期时长(h)": np.nan,
        "对数期时长(h)": np.nan,
        "稳定期时长(h)": np.nan,
    }


def _finite_or_neg_inf(value):
    return value if value is not None and np.isfinite(value) else -np.inf


def _to_float_array(series):
    return pd.to_numeric(series, errors="coerce").to_numpy(dtype=float)


def _safe_last_value(values, valid_num):
    if len(values) == 0:
        return np.nan
    index = min(max(valid_num - 1, 0), len(values) - 1)
    return values[index]


def _save_figure_to_buffer(fig):
    buffer = io.BytesIO()
    fig.savefig(buffer, format="png", dpi=300, bbox_inches="tight")
    buffer.seek(0)
    plt.close(fig)
    return buffer


def _style_time_axis(ax):
    ax.set_xlabel("Cultivation time (h)")
    ax.xaxis.set_major_locator(MultipleLocator(12))
    ax.xaxis.set_minor_locator(MultipleLocator(3))
    y_tick_interval = (ax.get_yticks()[1] - ax.get_yticks()[0]) if len(ax.get_yticks()) > 1 else 1
    if y_tick_interval == 0:
        y_tick_interval = 1
    ax.yaxis.set_minor_locator(MultipleLocator(y_tick_interval / 5))


def _insert_image(ws, buffer, anchor, width, height):
    img = Image(buffer)
    img.width = width
    img.height = height
    ws.add_image(img, anchor)


def _fit_logistic_curve(t, cell_counts, mu_guess=0.1, k_guess=None, n0_guess=None):
    t = np.asarray(t, dtype=float)
    cell_counts = np.asarray(cell_counts, dtype=float)

    if t.size < 3:
        raise ValueError("数据点不足，无法拟合")
    if not np.isfinite(t).all() or not np.isfinite(cell_counts).all():
        raise ValueError("数据中包含非数值")

    max_count = float(np.max(cell_counts)) if cell_counts.size else 1.0
    if k_guess is None:
        k_guess = max(max_count, 1.0)
    if n0_guess is None:
        positive_counts = cell_counts[cell_counts > 0]
        if cell_counts[0] > 0:
            n0_guess = float(cell_counts[0])
        elif positive_counts.size > 0:
            n0_guess = float(positive_counts[0])
        else:
            n0_guess = 1e-6

    mu_candidates = [mu_guess, max(mu_guess * 0.5, 1e-4), 0.2, 0.05]
    last_error = None
    upper_k = max(float(k_guess) * 1.5, 0.2)
    upper_n0 = max(max_count * 0.5 if max_count > 0 else 10.0, float(n0_guess) * 1.5, 1e-6)
    bounds = (
        [1e-6, 0.1, 1e-6],
        [1.0, upper_k, upper_n0],
    )

    for candidate_mu in mu_candidates:
        initial_guess = np.array([candidate_mu, k_guess, n0_guess], dtype=float)
        initial_guess = np.clip(initial_guess, np.array(bounds[0]) + 1e-12, np.array(bounds[1]) - 1e-12)
        try:
            popt, _ = curve_fit(
                f=modified_logistic_model,
                xdata=t,
                ydata=cell_counts,
                p0=initial_guess,
                bounds=bounds,
                maxfev=10000,
            )
            mu_max_fit, K_fit, N0_fit = popt
            y_fit = modified_logistic_model(t, mu_max_fit, K_fit, N0_fit)
            ss_res = float(np.sum((cell_counts - y_fit) ** 2))
            ss_tot = float(np.sum((cell_counts - np.mean(cell_counts)) ** 2))
            r2 = 1 - (ss_res / ss_tot) if ss_tot != 0 else np.nan
            t_d = np.log(2) / mu_max_fit if mu_max_fit != 0 else np.nan
            return {
                "mu_max": float(mu_max_fit),
                "K": float(K_fit),
                "N0": float(N0_fit),
                "r2": float(r2),
                "t_d": float(t_d),
                "y_fit": y_fit,
                "ss_res": ss_res,
                "ss_tot": ss_tot,
            }
        except Exception as exc:
            last_error = exc

    raise last_error if last_error is not None else RuntimeError("拟合失败")


def _build_subset_fit(
    selected_sheets,
    record_map,
    original_order,
    valid_num,
    fit_cache,
):
    key = tuple(sheet for sheet in original_order if sheet in selected_sheets)
    if key in fit_cache:
        return fit_cache[key]

    selected_records = [record_map[sheet] for sheet in original_order if sheet in selected_sheets]
    if len(selected_records) < 2:
        fit_cache[key] = None
        return None

    merged_t = np.concatenate([record["t"][:valid_num] for record in selected_records])
    merged_counts = np.concatenate([record["cell_counts"][:valid_num] for record in selected_records])

    try:
        n0_guess_values = [
            float(record["cell_counts"][0])
            for record in selected_records
            if len(record["cell_counts"]) > 0 and record["cell_counts"][0] > 0
        ]
        n0_guess = float(np.mean(n0_guess_values)) if n0_guess_values else None
        fit = _fit_logistic_curve(
            merged_t,
            merged_counts,
            mu_guess=0.1,
            k_guess=float(np.max(merged_counts)) if merged_counts.size else 1.0,
            n0_guess=n0_guess,
        )
    except Exception:
        fit_cache[key] = None
        return None

    sorted_t = np.sort(merged_t)
    phase_t = np.unique(np.concatenate([record["t"][:valid_num] for record in selected_records]))
    phase_duration, _ = calculate_growth_phases(phase_t, fit["mu_max"], fit["K"], fit["N0"])
    residuals = {
        record["sheet"]: float(
            np.sum(
                (
                    record["cell_counts"][:valid_num]
                    - modified_logistic_model(record["t"][:valid_num], fit["mu_max"], fit["K"], fit["N0"])
                )
                ** 2
            )
        )
        for record in selected_records
    }
    result = {
        "valid_sheets": [record["sheet"] for record in selected_records],
        "mu_max": fit["mu_max"],
        "K": fit["K"],
        "N0": fit["N0"],
        "r2": fit["r2"],
        "t_d": fit["t_d"],
        "merged_t": merged_t,
        "merged_counts": merged_counts,
        "sorted_t": sorted_t,
        "phase_t": phase_t,
        "phase_duration": phase_duration,
        "residuals": residuals,
        "y_fit": fit["y_fit"],
    }
    fit_cache[key] = result
    return result


def _repair_subset(active_set, removed_order, record_map, original_order, valid_num, fit_cache, target_r2):
    current_result = _build_subset_fit(active_set, record_map, original_order, valid_num, fit_cache)
    if current_result is None:
        return None, active_set, removed_order

    while removed_order:
        candidates = []
        for sheet in reversed(removed_order):
            trial_set = set(active_set)
            trial_set.add(sheet)
            trial_result = _build_subset_fit(trial_set, record_map, original_order, valid_num, fit_cache)
            if trial_result and trial_result["r2"] >= target_r2:
                candidates.append((sheet, trial_result))

        if not candidates:
            break

        chosen_sheet, chosen_result = max(candidates, key=lambda item: item[1]["r2"])
        active_set = set(chosen_result["valid_sheets"])
        removed_order = [sheet for sheet in removed_order if sheet != chosen_sheet]
        current_result = chosen_result
        print(
            f"回看修复：重新加入腔室 {chosen_sheet}，当前R²={current_result['r2']:.6f}，"
            f"保留{len(active_set)}个腔室"
        )

    return current_result, active_set, removed_order


def _select_best_subset(record_map, original_order, valid_num, summary_map, target_r2=0.8):
    fit_cache = {}
    active_set = set(original_order)
    removed_order = []
    best_seen_result = None

    current_result = _build_subset_fit(active_set, record_map, original_order, valid_num, fit_cache)
    if current_result is not None:
        best_seen_result = current_result
        if current_result["r2"] >= target_r2:
            return current_result, best_seen_result

    while len(active_set) >= 2:
        active_order = [sheet for sheet in original_order if sheet in active_set]
        candidate_results = []
        for sheet in active_order:
            trial_set = set(active_set)
            trial_set.remove(sheet)
            trial_result = _build_subset_fit(trial_set, record_map, original_order, valid_num, fit_cache)
            candidate_results.append((sheet, trial_result))

        feasible_candidates = [
            (sheet, trial_result)
            for sheet, trial_result in candidate_results
            if trial_result is not None and trial_result["r2"] >= target_r2
        ]

        if feasible_candidates:
            chosen_sheet, chosen_result = max(feasible_candidates, key=lambda item: item[1]["r2"])
        else:
            chosen_sheet, chosen_result = max(
                candidate_results,
                key=lambda item: _finite_or_neg_inf(item[1]["r2"]) if item[1] is not None else -np.inf,
            )

        if chosen_result is None:
            if current_result is not None and current_result["residuals"]:
                chosen_sheet = max(current_result["residuals"], key=current_result["residuals"].get)
            else:
                chosen_sheet = min(
                    active_order,
                    key=lambda sheet: _finite_or_neg_inf(summary_map.get(sheet, {}).get("拟合优度R²")),
                )
            trial_set = set(active_set)
            trial_set.remove(chosen_sheet)
            chosen_result = _build_subset_fit(trial_set, record_map, original_order, valid_num, fit_cache)

        active_set.remove(chosen_sheet)
        removed_order.append(chosen_sheet)
        current_result = chosen_result
        if current_result is not None:
            if (
                best_seen_result is None
                or _finite_or_neg_inf(current_result["r2"]) > _finite_or_neg_inf(best_seen_result["r2"])
            ):
                best_seen_result = current_result
            print(
                f"后向剔除：移除腔室 {chosen_sheet}，重拟合后R²={current_result['r2']:.6f}，"
                f"剩余{len(active_set)}个腔室"
            )
        else:
            print(f"后向剔除：移除腔室 {chosen_sheet} 后拟合失败，剩余{len(active_set)}个腔室")

        if current_result is not None and current_result["r2"] >= target_r2:
            repaired_result, repaired_set, repaired_removed = _repair_subset(
                active_set,
                removed_order,
                record_map,
                original_order,
                valid_num,
                fit_cache,
                target_r2,
            )
            return repaired_result, best_seen_result

    if current_result is not None and current_result["r2"] >= target_r2:
        repaired_result, _, _ = _repair_subset(
            active_set,
            removed_order,
            record_map,
            original_order,
            valid_num,
            fit_cache,
            target_r2,
        )
        return repaired_result, best_seen_result

    return None, best_seen_result


def _fit_and_write_individual_sheet(
    ws,
    sheet,
    plt_name,
    t,
    cell_counts,
    valid_num,
    df,
):
    fit_result = _fit_logistic_curve(t, cell_counts)
    phase_duration, _ = calculate_growth_phases(t, fit_result["mu_max"], fit_result["K"], fit_result["N0"])
    lag_duration = phase_duration["滞后期时长(h)"]
    log_duration = phase_duration["对数期时长(h)"]
    stable_duration = phase_duration["稳定期时长(h)"]
    F_last_cell_number = _safe_last_value(cell_counts, valid_num)
    growth_rate = (
        F_last_cell_number / valid_num
        if F_last_cell_number is not None and np.isfinite(F_last_cell_number)
        else np.nan
    )

    row = {
        "腔室名称": sheet,
        "最大比生长速率μmax (h^-1)": round(fit_result["mu_max"], 4),
        "环境容纳量K (个/腔室)": round(fit_result["K"], 2),
        "初始细胞数量N0 (个/腔室)": round(fit_result["N0"], 2),
        "拟合优度R²": round(fit_result["r2"], 4),
        "平均细胞周期T_d(h)": round(fit_result["t_d"], 2) if np.isfinite(fit_result["t_d"]) else np.nan,
        "增殖倍数 F": round(F_last_cell_number, 2) if np.isfinite(F_last_cell_number) else np.nan,
        "生长效率 η (1/增殖倍数)": round(growth_rate, 3) if np.isfinite(growth_rate) else np.nan,
        "滞后期时长(h)": lag_duration,
        "对数期时长(h)": log_duration,
        "稳定期时长(h)": stable_duration,
    }

    fig = plt.figure(figsize=(10, 6))
    plt.scatter(t, cell_counts, label="Actual data", color="blue", alpha=0.6)
    plt.plot(t, fit_result["y_fit"], label="Fitted curve", color="red", linewidth=2)
    ax = plt.gca()
    ax.set_xlabel("Cultivation time (h)")
    ax.set_ylabel("Relative cell number (normalized to 0 h)")
    ax.set_title(f"{plt_name} Cell growth curve fitting")
    _style_time_axis(ax)
    param_text = (
        f"$\\it{{μ}}_{{\\mathrm{{max}}}}$= {round(fit_result['mu_max'], 4)} h⁻¹\n"
        f"$\\it{{K}}$ = {round(fit_result['K'], 2)}\n"
        f"$\\it{{N}}_{{0}}$ = {round(fit_result['N0'], 2)}\n"
        f"R² = {round(fit_result['r2'], 4)}"
    )
    plt.text(
        0.05,
        0.95,
        param_text,
        transform=ax.transAxes,
        verticalalignment="top",
        fontsize=20,
        bbox=dict(boxstyle="round", facecolor="white", alpha=0.8),
    )
    plt.legend()
    plt.grid(alpha=0.3)
    img_buffer = _save_figure_to_buffer(fig)

    param_col_start = 5
    ws.cell(row=1, column=param_col_start).value = "参数名"
    ws.cell(row=1, column=param_col_start + 1).value = "参数值"
    ws.cell(row=2, column=param_col_start).value = "最大比生长速率uMAX(h^-1)"
    ws.cell(row=2, column=param_col_start + 1).value = round(fit_result["mu_max"], 4)
    ws.cell(row=3, column=param_col_start).value = "环境容纳量K(个)"
    ws.cell(row=3, column=param_col_start + 1).value = round(fit_result["K"], 4)
    ws.cell(row=4, column=param_col_start).value = "初始细胞数量N0(个)"
    ws.cell(row=4, column=param_col_start + 1).value = round(fit_result["N0"], 4)
    ws.cell(row=5, column=param_col_start).value = "拟合优度R²"
    ws.cell(row=5, column=param_col_start + 1).value = round(fit_result["r2"], 4)
    ws.cell(row=6, column=param_col_start).value = "平均细胞周期T_d(h)"
    ws.cell(row=6, column=param_col_start + 1).value = round(fit_result["t_d"], 2) if np.isfinite(fit_result["t_d"]) else np.nan
    ws.cell(row=7, column=param_col_start).value = "增殖倍数 F"
    ws.cell(row=7, column=param_col_start + 1).value = round(F_last_cell_number, 2) if np.isfinite(F_last_cell_number) else np.nan
    ws.cell(row=8, column=param_col_start).value = "生长效率 η (1/增殖倍数)"
    ws.cell(row=8, column=param_col_start + 1).value = round(growth_rate, 3) if np.isfinite(growth_rate) else np.nan
    ws.cell(row=9, column=param_col_start).value = "滞后期时长(h)"
    ws.cell(row=9, column=param_col_start + 1).value = lag_duration
    ws.cell(row=10, column=param_col_start).value = "对数期时长(h)"
    ws.cell(row=10, column=param_col_start + 1).value = log_duration
    ws.cell(row=11, column=param_col_start).value = "稳定期时长(h)"
    ws.cell(row=11, column=param_col_start + 1).value = stable_duration
    _insert_image(ws, img_buffer, "H02", 600, 400)

    return row


def _result_priority(result):
    if result is None:
        return (-np.inf, -np.inf, -np.inf, -np.inf)
    residuals = result.get("residuals") or {}
    r2 = _finite_or_neg_inf(result.get("r2"))
    ss_res = result.get("ss_res", np.inf)
    total_residual = -float(ss_res) if np.isfinite(ss_res) else -np.inf
    max_residual = -float(max(residuals.values())) if residuals else -np.inf
    count = len(result.get("valid_sheets", []))
    return (r2, count, total_residual, max_residual)


def _state_priority(state):
    return _result_priority(state.get("result"))


def _rank_removal_candidates(state, active_order, summary_map, branch_limit):
    branch_limit = max(1, min(branch_limit, len(active_order)))
    result = state.get("result")
    if result is not None and result.get("residuals"):
        ordered = sorted(
            active_order,
            key=lambda sheet: (
                _finite_or_neg_inf(result["residuals"].get(sheet)),
                -_finite_or_neg_inf(summary_map.get(sheet, {}).get("拟合优度R²")),
                sheet,
            ),
            reverse=True,
        )
    else:
        ordered = sorted(
            active_order,
            key=lambda sheet: (
                _finite_or_neg_inf(summary_map.get(sheet, {}).get("拟合优度R²")),
                sheet,
            ),
        )
    return ordered[:branch_limit]


def _select_best_subset_beam(
    record_map,
    original_order,
    valid_num,
    summary_map,
    target_r2=0.8,
    beam_width=None,
    branch_limit=None,
):
    fit_cache = {}
    if beam_width is None:
        beam_width = min(10, max(4, int(np.ceil(len(original_order) / 6))))
    if branch_limit is None:
        branch_limit = min(10, max(4, int(np.ceil(len(original_order) / 6))))

    active_set = set(original_order)
    start_result = _build_subset_fit(active_set, record_map, original_order, valid_num, fit_cache)
    start_state = {
        "active_set": active_set,
        "removed_order": [],
        "result": start_result,
    }
    best_seen_result = start_result

    if start_result is not None and start_result["r2"] >= target_r2:
        repaired_result, _, _ = _repair_subset(
            active_set,
            [],
            record_map,
            original_order,
            valid_num,
            fit_cache,
            target_r2,
        )
        final_result = repaired_result or start_result
        return final_result, final_result

    current_layer = [start_state]
    depth = 0

    while current_layer:
        next_states = {}
        for state in current_layer:
            active_order = [sheet for sheet in original_order if sheet in state["active_set"]]
            if len(active_order) <= 1:
                continue

            removal_candidates = _rank_removal_candidates(state, active_order, summary_map, branch_limit)
            for sheet in removal_candidates:
                trial_set = set(state["active_set"])
                trial_set.remove(sheet)
                trial_result = _build_subset_fit(trial_set, record_map, original_order, valid_num, fit_cache)
                if trial_result is None:
                    continue

                trial_state = {
                    "active_set": trial_set,
                    "removed_order": state["removed_order"] + [sheet],
                    "result": trial_result,
                }
                key = tuple(trial_result["valid_sheets"])
                existing = next_states.get(key)
                if existing is None or _state_priority(trial_state) > _state_priority(existing):
                    next_states[key] = trial_state

                if best_seen_result is None or _result_priority(trial_result) > _result_priority(best_seen_result):
                    best_seen_result = trial_result

        if not next_states:
            break

        candidate_states = list(next_states.values())
        feasible_states = [
            state for state in candidate_states
            if state["result"] is not None and state["result"]["r2"] >= target_r2
        ]
        if feasible_states:
            chosen_state = max(feasible_states, key=_state_priority)
            repaired_result, _, _ = _repair_subset(
                chosen_state["active_set"],
                chosen_state["removed_order"],
                record_map,
                original_order,
                valid_num,
                fit_cache,
                target_r2,
            )
            final_result = repaired_result or chosen_state["result"]
            if final_result is not None and (
                best_seen_result is None
                or _result_priority(final_result) > _result_priority(best_seen_result)
            ):
                best_seen_result = final_result
            print(
                f"Beam search depth {depth + 1}: selected {len(final_result['valid_sheets']) if final_result else 0} "
                f"chambers, R²={final_result['r2']:.6f}" if final_result else
                f"Beam search depth {depth + 1}: selected a feasible subset"
            )
            return final_result, best_seen_result

        current_layer = sorted(candidate_states, key=_state_priority, reverse=True)[:beam_width]
        depth += 1
        print(f"Beam search depth {depth}: kept {len(current_layer)} candidate subsets")

    return None, best_seen_result


def process_cell_growth(
    excel_path,
    result_excel_path="enhanced_cell_growth_results.xlsx",
    min_data_points=5,
    nitrogen_concentration="test",
    skip_sheet_name={"数据汇总"},
    valid_num=96,
):
    skip_sheet = skip_sheet_name
    wb_original = openpyxl.load_workbook(excel_path)
    excel_file = pd.ExcelFile(excel_path)

    try:
        sheet_names = wb_original.sheetnames
        individual_summary = []
        individual_data = []
        record_map = {}
        summary_map = {}
        eligible_order = []
        merged_params = None
        merged_img_buffer = None
        target_r2 = 0.8

        for sheet in sheet_names:
            if sheet in skip_sheet:
                print(f"跳过页签: {sheet}")
                continue

            print(f"正在处理页签: {sheet}")
            plt_name = nitrogen_concentration + "_" + sheet[4:]
            ws = wb_original[sheet]
            try:
                df = excel_file.parse(sheet_name=sheet, nrows=valid_num + 1)
            except Exception as e:
                print(f"页签 {sheet} 读取失败：{str(e)}\n")
                row = _empty_summary_row(sheet)
                individual_summary.append(row)
                summary_map[sheet] = row
                continue

            required_columns = ["目标数量", "总面积(μm²)", "相对平均细胞面积"]
            missing_cols = [col for col in required_columns if col not in df.columns]
            if missing_cols:
                print(f"页签 {sheet} 缺少必要列: {missing_cols}，跳过\n")
                row = _empty_summary_row(sheet)
                individual_summary.append(row)
                summary_map[sheet] = row
                continue

            cell_counts = _to_float_array(df["目标数量"])
            if cell_counts.size == 0 or np.all(np.nan_to_num(cell_counts, nan=0.0) == 0):
                print(f"页签 {sheet} 所有细胞数量为0，跳过\n")
                row = _empty_summary_row(sheet)
                individual_summary.append(row)
                summary_map[sheet] = row
                continue

            t = np.arange(1, len(cell_counts) + 1)
            record = {
                "sheet": sheet,
                "t": t,
                "cell_counts": cell_counts,
                "area": _to_float_array(df["总面积(μm²)"]),
                "avg_cell_area": _to_float_array(df["相对平均细胞面积"]),
            }

            if len(cell_counts) >= min_data_points:
                eligible_order.append(sheet)
                individual_data.append(record)
                record_map[sheet] = record

            try:
                row = _fit_and_write_individual_sheet(
                    ws=ws,
                    sheet=sheet,
                    plt_name=plt_name,
                    t=t,
                    cell_counts=cell_counts,
                    valid_num=valid_num,
                    df=df,
                )
                individual_summary.append(row)
                summary_map[sheet] = row
                print(f"页签 {sheet} 单独拟合完成\n")
            except Exception as e:
                print(f"页签 {sheet} 单独拟合失败：{str(e)}\n")
                row = _empty_summary_row(sheet)
                individual_summary.append(row)
                summary_map[sheet] = row

        if len(eligible_order) >= 2:
            merged_params, best_merged_result = _select_best_subset_beam(
                record_map=record_map,
                original_order=eligible_order,
                valid_num=valid_num,
                summary_map=summary_map,
                target_r2=target_r2,
            )
            if merged_params is not None:
                merged_img_buffer = None
                print(
                    f"合并拟合完成（已后向剔除并回看修复）："
                    f"μmax={merged_params['mu_max']:.6f}, "
                    f"K={merged_params['K']:.6f}, "
                    f"N0={merged_params['N0']:.6f}, "
                    f"R²={merged_params['r2']:.6f}"
                )

                fig = plt.figure(figsize=(12, 7))
                plt.scatter(
                    merged_params["merged_t"],
                    merged_params["merged_counts"],
                    label="Merged actual data",
                    color="blue",
                    alpha=0.5,
                    s=30,
                )
                sorted_t = merged_params["sorted_t"]
                plt.plot(
                    sorted_t,
                    modified_logistic_model(
                        sorted_t,
                        merged_params["mu_max"],
                        merged_params["K"],
                        merged_params["N0"],
                    ),
                    label="Merged fitted curve",
                    color="#F2BA02",
                    linewidth=2,
                )
                ax = plt.gca()
                ax.set_xlabel("Cultivation time (h)")
                ax.set_ylabel("Relative cell number (normalized to 0 h)")
                ax.set_title(
                    f"Growth curve (merged data, {len(merged_params['valid_sheets'])} chambers, "
                    f"NH$_4^+$-N={nitrogen_concentration})"
                )
                _style_time_axis(ax)
                param_text = (
                    f"$\\it{{μ}}_{{\\mathrm{{max}}}}$ = {merged_params['mu_max']:.4f} h⁻¹\n"
                    f"$\\it{{K}}$ = {merged_params['K']:.2f}\n"
                    f"$\\it{{N}}_{{0}}$ = {merged_params['N0']:.2f}\n"
                    f"R² = {merged_params['r2']:.4f}"
                )
                plt.text(
                    0.05,
                    0.95,
                    param_text,
                    transform=ax.transAxes,
                    verticalalignment="top",
                    fontsize=20,
                    bbox=dict(boxstyle="round", facecolor="white", alpha=0.8),
                )
                plt.legend()
                plt.grid(alpha=0.3)
                merged_img_buffer = _save_figure_to_buffer(fig)
                merged_params["phase_duration"], _ = calculate_growth_phases(
                    merged_params["phase_t"],
                    merged_params["mu_max"],
                    merged_params["K"],
                    merged_params["N0"],
                )
            else:
                if best_merged_result is not None:
                    print(
                        f"无法通过后向剔除与回看修复使R²≥{target_r2}，"
                        f"最佳结果为R²={best_merged_result['r2']:.2f}，"
                        f"保留{len(best_merged_result['valid_sheets'])}个腔室"
                    )
                else:
                    print("无法得到可用的合并拟合结果")

        if "汇总结果" in wb_original.sheetnames:
            del wb_original["汇总结果"]
        summary_ws = wb_original.create_sheet(title="汇总结果", index=0)
        summary_ws["A1"] = "腔室名称"
        summary_ws["B1"] = "最大比生长速率μmax (h^-1)"
        summary_ws["C1"] = "环境容纳量K (个/腔室)"
        summary_ws["D1"] = "初始细胞数量N0 (个/腔室)"
        summary_ws["E1"] = "拟合优度R²"
        summary_ws["F1"] = "平均细胞周期T_d(h)"
        summary_ws["G1"] = "增殖倍数 F"
        summary_ws["H1"] = "生长效率 η (1/增殖倍数)"
        summary_ws["I1"] = "滞后期时长(h)"
        summary_ws["J1"] = "对数期时长(h)"
        summary_ws["K1"] = "稳定期时长(h)"

        selected_sheets = merged_params.get("valid_sheets", []) if merged_params else []
        valid_data = [item for item in individual_summary if item["腔室名称"] in selected_sheets]
        non_valid_data = [item for item in individual_summary if item["腔室名称"] not in selected_sheets]

        current_row = 2
        for data in valid_data:
            summary_ws[f"A{current_row}"] = data["腔室名称"]
            summary_ws[f"B{current_row}"] = data["最大比生长速率μmax (h^-1)"]
            summary_ws[f"C{current_row}"] = data["环境容纳量K (个/腔室)"]
            summary_ws[f"D{current_row}"] = data["初始细胞数量N0 (个/腔室)"]
            summary_ws[f"E{current_row}"] = data["拟合优度R²"]
            summary_ws[f"F{current_row}"] = data["平均细胞周期T_d(h)"]
            summary_ws[f"G{current_row}"] = data["增殖倍数 F"]
            summary_ws[f"H{current_row}"] = data["生长效率 η (1/增殖倍数)"]
            summary_ws[f"I{current_row}"] = data["滞后期时长(h)"]
            summary_ws[f"J{current_row}"] = data["对数期时长(h)"]
            summary_ws[f"K{current_row}"] = data["稳定期时长(h)"]
            current_row += 1

        if non_valid_data:
            summary_ws[f"A{current_row}"] = "未参与拟合数据"
            summary_ws.merge_cells(f"A{current_row}:K{current_row}")
            summary_ws[f"A{current_row}"].font = openpyxl.styles.Font(bold=True, color="FF0000")
            current_row += 1

        for data in non_valid_data:
            summary_ws[f"A{current_row}"] = data["腔室名称"]
            summary_ws[f"B{current_row}"] = data["最大比生长速率μmax (h^-1)"]
            summary_ws[f"C{current_row}"] = data["环境容纳量K (个/腔室)"]
            summary_ws[f"D{current_row}"] = data["初始细胞数量N0 (个/腔室)"]
            summary_ws[f"E{current_row}"] = data["拟合优度R²"]
            summary_ws[f"F{current_row}"] = data["平均细胞周期T_d(h)"]
            summary_ws[f"G{current_row}"] = data["增殖倍数 F"]
            summary_ws[f"H{current_row}"] = data["生长效率 η (1/增殖倍数)"]
            summary_ws[f"I{current_row}"] = data["滞后期时长(h)"]
            summary_ws[f"J{current_row}"] = data["对数期时长(h)"]
            summary_ws[f"K{current_row}"] = data["稳定期时长(h)"]
            current_row += 1

        for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            summary_ws.column_dimensions[col].width = 25
        for col in ["I", "J", "K"]:
            summary_ws.column_dimensions[col].width = 20

        reserve_row = current_row + 1

        if merged_params is not None:
            last_row = current_row + 1
            summary_ws[f"A{last_row}"] = "合并拟合结果"
            summary_ws[f"A{last_row}"].font = openpyxl.styles.Font(bold=True)
            merged_phase_duration = merged_params["phase_duration"]
            param_rows = {
                "参与拟合的页签数量": len(merged_params["valid_sheets"]),
                "总数据点数量": len(merged_params["merged_t"]),
                "最大比生长速率μmax (h^-1)": round(merged_params["mu_max"], 4),
                "环境容纳量K (个/腔室)": round(merged_params["K"], 2),
                "初始细胞数量N0 (个/腔室)": round(merged_params["N0"], 2),
                "拟合优度R²": round(merged_params["r2"], 4),
                "滞后期时长(h)": merged_phase_duration["滞后期时长(h)"],
                "对数期时长(h)": merged_phase_duration["对数期时长(h)"],
                "稳定期时长(h)": merged_phase_duration["稳定期时长(h)"],
            }
            current_row = last_row + 1
            for param, value in param_rows.items():
                summary_ws[f"A{current_row}"] = param
                summary_ws[f"B{current_row}"] = value
                current_row += 1
            if merged_img_buffer is not None:
                summary_ws[f"C{last_row}"] = "参与拟合的腔室"
                summary_ws[f"D{last_row}"] = ", ".join(merged_params["valid_sheets"])
                _insert_image(summary_ws, merged_img_buffer, f"C{last_row + 1}", 700, 500)

        if merged_params is not None and len(selected_sheets) > 0:
            for sheet_data in individual_data:
                sheet_name = sheet_data["sheet"]
                plt_name = nitrogen_concentration + "_" + sheet_name[4:]
                t_data = sheet_data["t"]
                counts_data = sheet_data["cell_counts"]
                fig = plt.figure(figsize=(10, 6))
                plt.scatter(t_data, counts_data, label=f"{plt_name} actual data", color="blue", alpha=0.6)
                merged_curve = modified_logistic_model(
                    t_data,
                    merged_params["mu_max"],
                    merged_params["K"],
                    merged_params["N0"],
                )
                plt.plot(t_data, merged_curve, label="Merged fitted curve", color="#F2BA02", linewidth=2)
                ax = plt.gca()
                ax.set_xlabel("Cultivation time (h)")
                ax.set_ylabel("Relative cell number (normalized to 0 h)")
                ax.set_title(f"{plt_name} data vs merged fitting curve")
                _style_time_axis(ax)
                param_text = (
                    f"$\\it{{μ}}_{{\\mathrm{{max}}}}$ = {merged_params['mu_max']:.4f} h⁻¹\n"
                    f"$\\it{{K}}$ = {merged_params['K']:.2f}\n"
                    f"$\\it{{N}}_{{0}}$ = {merged_params['N0']:.2f}\n"
                    f"R² = {merged_params['r2']:.4f}"
                )
                plt.text(
                    0.05,
                    0.95,
                    param_text,
                    transform=ax.transAxes,
                    verticalalignment="top",
                    fontsize=20,
                    bbox=dict(boxstyle="round", facecolor="white", alpha=0.8),
                )
                plt.legend()
                plt.grid(alpha=0.3)
                img_buffer = _save_figure_to_buffer(fig)
                try:
                    ws = wb_original[sheet_name]
                    _insert_image(ws, img_buffer, "H24", 600, 400)
                    print(f"已为 {sheet_name} 添加合并拟合对比图")
                except Exception as e:
                    print(f"为 {sheet_name} 添加对比图失败：{str(e)}")

        for sheet_data in individual_data:
            sheet_name = sheet_data["sheet"]
            t_data = sheet_data["t"][:valid_num]
            area = sheet_data["area"][:valid_num]
            avg_cell_area = sheet_data["avg_cell_area"][:valid_num]
            try:
                ws = wb_original[sheet_name]
                plt_title = nitrogen_concentration + "_" + sheet_name[4:]

                fig = plt.figure(figsize=(10, 6))
                plt.scatter(t_data, area, label="Relative total cell area", color="green", alpha=0.6)
                ax = plt.gca()
                ax.set_xlabel("Cultivation time (h)")
                ax.set_ylabel("Relative total cell area (normalized to 0 h)")
                ax.set_title(f"{plt_title} - Relative total cell area variation")
                _style_time_axis(ax)
                plt.legend()
                plt.grid(alpha=0.3)
                area_buffer = _save_figure_to_buffer(fig)
                _insert_image(ws, area_buffer, "R2", 600, 400)

                fig = plt.figure(figsize=(10, 6))
                plt.scatter(t_data, avg_cell_area, label="Relative average cell area", color="purple", alpha=0.6, s=30)
                plt.plot(t_data, avg_cell_area, color="purple", alpha=0.8, linestyle="-", linewidth=1.5, marker="", markersize=5)
                ax = plt.gca()
                ax.set_xlabel("Cultivation time (h)")
                ax.set_ylabel("Relative average cell area (normalized to 0 h)")
                ax.set_title(f"{plt_title} - Relative average cell area variation")
                _style_time_axis(ax)
                plt.legend()
                plt.grid(alpha=0.3)
                avg_buffer = _save_figure_to_buffer(fig)
                _insert_image(ws, avg_buffer, "R24", 600, 400)
                print(f"已为 {sheet_name} 添加2个图表")
            except Exception as e:
                print(f"为 {sheet_name} 添加趋势图表失败：{str(e)}")

        if merged_params is not None and len(selected_sheets) > 0:
            all_t = []
            all_area = []
            all_avg_area = []
            for sheet_data in individual_data:
                if sheet_data["sheet"] in merged_params["valid_sheets"]:
                    all_t.extend(sheet_data["t"][:valid_num])
                    all_area.extend(sheet_data["area"][:valid_num])
                    all_avg_area.extend(sheet_data["avg_cell_area"][:valid_num])

            fig = plt.figure(figsize=(12, 7))
            plt.scatter(all_t, all_area, label="Relative total area data", color="green", alpha=0.6, s=30)
            ax = plt.gca()
            ax.set_xlabel("Cultivation time (h)")
            ax.set_ylabel("Relative total cell area (normalized to 0 h)")
            ax.set_title(
                f"Relative total area (merged data, {len(merged_params['valid_sheets'])} chambers, "
                f"NH$_4^+$-N={nitrogen_concentration})"
            )
            _style_time_axis(ax)
            plt.legend()
            plt.grid(alpha=0.3)
            total_area_buffer = _save_figure_to_buffer(fig)

            fig = plt.figure(figsize=(12, 7))
            plt.scatter(all_t, all_avg_area, label="Relative average cell area data", color="purple", alpha=0.6, s=30)
            ax = plt.gca()
            ax.set_xlabel("Cultivation time (h)")
            ax.set_ylabel("Relative average cell area (normalized to 0 h)")
            ax.set_title(
                f"Relative average cell area (merged data, {len(merged_params['valid_sheets'])} chambers, "
                f"NH$_4^+$-N={nitrogen_concentration})"
            )
            _style_time_axis(ax)
            plt.legend()
            plt.grid(alpha=0.3)
            avg_area_buffer = _save_figure_to_buffer(fig)

            current_row = reserve_row + 1
            _insert_image(summary_ws, total_area_buffer, f"G{current_row}", 700, 500)
            _insert_image(summary_ws, avg_area_buffer, f"L{current_row}", 700, 500)
            print("已在汇总结果页签添加2个汇总散点图")

        for sheet_data in individual_data:
            sheet_name = sheet_data["sheet"]
            try:
                ws = wb_original[sheet_name]
                title_row = 1
                area_col = None
                avg_area_col = None
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=title_row, column=col).value
                    if cell_value == "总面积(μm²)":
                        area_col = col
                    elif cell_value == "相对平均细胞面积":
                        avg_area_col = col
                    if area_col and avg_area_col:
                        break

                if area_col:
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=area_col)
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.value = round(cell.value, 2)
                            cell.number_format = numbers.FORMAT_NUMBER_00

                if avg_area_col:
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=avg_area_col)
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.value = round(cell.value, 2)
                            cell.number_format = numbers.FORMAT_NUMBER_00

                print(f"已为 {sheet_name} 格式化面积数据为两位小数")
            except Exception as e:
                print(f"格式化 {sheet_name} 面积数据时出错：{str(e)}")

        wb_original.save(result_excel_path)
        print(f"所有处理完成，结果已保存到：{result_excel_path}")
        return pd.DataFrame(individual_summary)
    finally:
        try:
            excel_file.close()
        except Exception:
            pass
        wb_original.close()


# ---------------------- 脚本执行入口 ----------------------
if __name__ == "__main__":
    EXCEL_FILE_PATH = r"F:\Microalgae_Photoes\20260504\数据汇总\02_标准化数据\CH3_标准化.xlsx"
    RESULT_EXCEL_PATH = r"F:\Microalgae_Photoes\20260504\数据汇总\03_可视化结果\CH3_可视化结果.xlsx"
    Nitrogen = "160 mg/L"
    skip_sheet = {"数据汇总", "👈←25-10-25批次   理论预测17组→👉", "← 👈25-09-26批  理论预测2组👉→"}
    set_valid_num = 97
    result = process_cell_growth(
        excel_path=EXCEL_FILE_PATH,
        result_excel_path=RESULT_EXCEL_PATH,
        min_data_points=5,
        nitrogen_concentration=Nitrogen,
        skip_sheet_name=skip_sheet,
        valid_num=set_valid_num,
    )
    if result is not None:
        print("\n单独拟合结果预览（含生长阶段时长）：")
        print(result[["腔室名称", "滞后期时长(h)", "对数期时长(h)", "稳定期时长(h)"]].head())
